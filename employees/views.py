from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Q
from .models import Position, Employee
from .forms import PositionForm, EmployeeForm

from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.core.paginator import Paginator

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime

# --- TOGGLE VIEW MODE ACTION ---

def toggle_view_mode(request):
    mode = request.GET.get('mode', 'fbv')
    if mode in ['fbv', 'cbv']:
        request.session['view_mode'] = mode
    referer = request.META.get('HTTP_REFERER')
    return redirect(referer or 'home')


# ==============================================================================
# 1. FUNCTION-BASED VIEWS (FBV) IMPLEMENTATIONS
# ==============================================================================

@login_required
def home_view_fbv(request):
    total_employees = Employee.objects.filter(is_active=True).count()
    total_positions = Position.objects.count()
    total_payroll = Employee.objects.filter(is_active=True).aggregate(total=Sum('salary'))['total'] or 0.00
    recent_employees = Employee.objects.select_related('position').order_by('-hire_date', '-id')[:5]
    
    context = {
        'total_employees': total_employees,
        'total_positions': total_positions,
        'total_payroll': total_payroll,
        'recent_employees': recent_employees,
    }
    return render(request, 'employees/home.html', context)


@login_required
def position_list_fbv(request):
    positions = Position.objects.all().order_by('name')
    q = request.GET.get('q')
    field = request.GET.get('field')
    
    if q and field:
        if field == 'name':
            positions = positions.filter(name__icontains=q)
        elif field == 'description':
            positions = positions.filter(description__icontains=q)
        elif field == 'all':
            positions = positions.filter(Q(name__icontains=q) | Q(description__icontains=q))
            
    paginator = Paginator(positions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'employees/position_list.html', {'positions': page_obj, 'page_obj': page_obj})


@login_required
def position_create_fbv(request):
    form = PositionForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('position_list')
    return render(request, 'employees/position_form.html', {'form': form, 'title': 'Nuevo Cargo'})


@login_required
def position_edit_fbv(request, pk):
    position = get_object_or_404(Position, pk=pk)
    form = PositionForm(request.POST or None, instance=position)
    if form.is_valid():
        form.save()
        return redirect('position_list')
    return render(request, 'employees/position_form.html', {'form': form, 'title': 'Editar Cargo'})


@login_required
def position_detail_fbv(request, pk):
    position = get_object_or_404(Position, pk=pk)
    return render(request, 'employees/position_detail.html', {'position': position})


@login_required
def position_delete_fbv(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == 'POST':
        position.delete()
        return redirect('position_list')
    return render(request, 'employees/position_delete_confirm.html', {'position': position})


@login_required
def employee_list_fbv(request):
    employees = Employee.objects.select_related('position').all().order_by('last_name')
    q = request.GET.get('q')
    field = request.GET.get('field')
    
    if q and field:
        if field == 'first_name':
            employees = employees.filter(first_name__icontains=q)
        elif field == 'last_name':
            employees = employees.filter(last_name__icontains=q)
        elif field == 'position':
            employees = employees.filter(position__name__icontains=q)
        elif field == 'status':
            is_active_val = None
            q_clean = q.lower().strip()
            if q_clean in ['activo', 'activa', 'si', 'sí', 'true', '1', 'active']:
                is_active_val = True
            elif q_clean in ['inactivo', 'inactiva', 'no', 'false', '0', 'inactive']:
                is_active_val = False
            
            if is_active_val is not None:
                employees = employees.filter(is_active=is_active_val)
        elif field == 'all':
            employees = employees.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(position__name__icontains=q)
            )
            
    paginator = Paginator(employees, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'employees/employee_list.html', {'employees': page_obj, 'page_obj': page_obj})


@login_required
def employee_create_fbv(request):
    form = EmployeeForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('employee_list')
    return render(request, 'employees/employee_form.html', {'form': form, 'title': 'Nuevo Empleado'})


@login_required
def employee_edit_fbv(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    form = EmployeeForm(request.POST or None, instance=employee)
    if form.is_valid():
        form.save()
        return redirect('employee_list')
    return render(request, 'employees/employee_form.html', {'form': form, 'title': 'Editar Empleado'})


@login_required
def employee_detail_fbv(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'employees/employee_detail.html', {'employee': employee})


@login_required
def employee_delete_fbv(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('employee_list')
    return render(request, 'employees/employee_delete_confirm.html', {'employee': employee})


@login_required
def employee_report_pdf_fbv(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_empleados.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20,
        textColor=colors.HexColor('#1e3c72'), spaceAfter=15, alignment=1
    )
    header_style = ParagraphStyle(
        'HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, textColor=colors.white
    )
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#333333')
    )
    
    story.append(Paragraph("Reporte General de Empleados", title_style))
    story.append(Paragraph(f"Generado el: {datetime.date.today().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 15))
    
    headers = [
        Paragraph("Nombres", header_style), Paragraph("Apellidos", header_style), Paragraph("Correo", header_style),
        Paragraph("Sueldo", header_style), Paragraph("Fecha Ingreso", header_style), Paragraph("Cargo", header_style),
        Paragraph("Estado", header_style)
    ]
    data = [headers]
    
    employees = Employee.objects.select_related('position').all().order_by('last_name')
    for emp in employees:
        data.append([
            Paragraph(emp.first_name, cell_style),
            Paragraph(emp.last_name, cell_style),
            Paragraph(emp.email, cell_style),
            Paragraph(f"${emp.salary}", cell_style),
            Paragraph(emp.hire_date.strftime('%d/%m/%Y') if emp.hire_date else "", cell_style),
            Paragraph(emp.position.name if emp.position else "", cell_style),
            Paragraph("Activo" if emp.is_active else "Inactivo", cell_style)
        ])
        
    t = Table(data, colWidths=[80, 80, 130, 60, 70, 70, 60])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3c72')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 6),
    ]))
    story.append(t)
    doc.build(story)
    return response


@login_required
def employee_report_excel_fbv(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )
    
    headers = ["Nombres", "Apellidos", "Correo", "Sueldo", "Fecha Ingreso", "Cargo", "Estado"]
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    employees = Employee.objects.select_related('position').all().order_by('last_name')
    for emp in employees:
        row_data = [
            emp.first_name, emp.last_name, emp.email,
            float(emp.salary) if emp.salary else 0.0,
            emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else "",
            emp.position.name if emp.position else "",
            "Activo" if emp.is_active else "Inactivo"
        ]
        ws.append(row_data)
        
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = left_align
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=4)
        cell.number_format = '"$"#,##0.00'
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_empleados.xlsx"'
    wb.save(response)
    return response


@login_required
def position_report_pdf_fbv(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_cargos.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20,
        textColor=colors.HexColor('#1e3c72'), spaceAfter=15, alignment=1
    )
    header_style = ParagraphStyle(
        'HeaderStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, textColor=colors.white
    )
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#333333')
    )
    
    story.append(Paragraph("Reporte de Cargos Disponibles", title_style))
    story.append(Paragraph(f"Generado el: {datetime.date.today().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 15))
    
    headers = [
        Paragraph("Nombre del Cargo", header_style), Paragraph("Descripción", header_style),
        Paragraph("Salario Máximo", header_style)
    ]
    data = [headers]
    
    positions = Position.objects.all().order_by('name')
    for pos in positions:
        data.append([
            Paragraph(pos.name, cell_style),
            Paragraph(pos.description if pos.description else "Sin descripción", cell_style),
            Paragraph(f"${pos.max_salary}" if pos.max_salary else "Sin límite", cell_style)
        ])
        
    t = Table(data, colWidths=[150, 250, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3c72')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 6),
    ]))
    story.append(t)
    doc.build(story)
    return response


@login_required
def position_report_excel_fbv(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cargos"
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )
    
    headers = ["Nombre del Cargo", "Descripción", "Salario Máximo"]
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    positions = Position.objects.all().order_by('name')
    for pos in positions:
        row_data = [
            pos.name, pos.description if pos.description else "",
            float(pos.max_salary) if pos.max_salary else ""
        ]
        ws.append(row_data)
        
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = left_align
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value != "":
            cell.number_format = '"$"#,##0.00'
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_cargos.xlsx"'
    wb.save(response)
    return response


# ==============================================================================
# 2. CLASS-BASED VIEWS (CBV) IMPLEMENTATIONS
# ==============================================================================

class HomeCBV(LoginRequiredMixin, TemplateView):
    template_name = 'employees/home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_employees'] = Employee.objects.filter(is_active=True).count()
        context['total_positions'] = Position.objects.count()
        context['total_payroll'] = Employee.objects.filter(is_active=True).aggregate(total=Sum('salary'))['total'] or 0.00
        context['recent_employees'] = Employee.objects.select_related('position').order_by('-hire_date', '-id')[:5]
        return context


class PositionListCBV(LoginRequiredMixin, ListView):
    model = Position
    template_name = 'employees/position_list.html'
    context_object_name = 'positions'
    paginate_by = 10
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get('q')
        field = self.request.GET.get('field')
        if q and field:
            if field == 'name':
                queryset = queryset.filter(name__icontains=q)
            elif field == 'description':
                queryset = queryset.filter(description__icontains=q)
            elif field == 'all':
                queryset = queryset.filter(Q(name__icontains=q) | Q(description__icontains=q))
        return queryset


class PositionCreateCBV(LoginRequiredMixin, CreateView):
    model = Position
    form_class = PositionForm
    template_name = 'employees/position_form.html'
    success_url = reverse_lazy('position_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuevo Cargo'
        return context


class PositionEditCBV(LoginRequiredMixin, UpdateView):
    model = Position
    form_class = PositionForm
    template_name = 'employees/position_form.html'
    success_url = reverse_lazy('position_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Cargo'
        return context


class PositionDetailCBV(LoginRequiredMixin, DetailView):
    model = Position
    template_name = 'employees/position_detail.html'
    context_object_name = 'position'


class PositionDeleteCBV(LoginRequiredMixin, DeleteView):
    model = Position
    template_name = 'employees/position_delete_confirm.html'
    success_url = reverse_lazy('position_list')
    context_object_name = 'position'


class EmployeeListCBV(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'employees/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 10
    ordering = ['last_name']

    def get_queryset(self):
        queryset = super().get_queryset().select_related('position')
        q = self.request.GET.get('q')
        field = self.request.GET.get('field')
        if q and field:
            if field == 'first_name':
                queryset = queryset.filter(first_name__icontains=q)
            elif field == 'last_name':
                queryset = queryset.filter(last_name__icontains=q)
            elif field == 'position':
                queryset = queryset.filter(position__name__icontains=q)
            elif field == 'status':
                is_active_val = None
                q_clean = q.lower().strip()
                if q_clean in ['activo', 'activa', 'si', 'sí', 'true', '1', 'active']:
                    is_active_val = True
                elif q_clean in ['inactivo', 'inactiva', 'no', 'false', '0', 'inactive']:
                    is_active_val = False
                if is_active_val is not None:
                    queryset = queryset.filter(is_active=is_active_val)
            elif field == 'all':
                queryset = queryset.filter(
                    Q(first_name__icontains=q) |
                    Q(last_name__icontains=q) |
                    Q(position__name__icontains=q)
                )
        return queryset


class EmployeeCreateCBV(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form.html'
    success_url = reverse_lazy('employee_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Nuevo Empleado'
        return context


class EmployeeEditCBV(LoginRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form.html'
    success_url = reverse_lazy('employee_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Editar Empleado'
        return context


class EmployeeDetailCBV(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = 'employees/employee_detail.html'
    context_object_name = 'employee'


class EmployeeDeleteCBV(LoginRequiredMixin, DeleteView):
    model = Employee
    template_name = 'employees/employee_delete_confirm.html'
    success_url = reverse_lazy('employee_list')
    context_object_name = 'employee'


class EmployeeReportPDFCBV(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return employee_report_pdf_fbv(request)


class EmployeeReportExcelCBV(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return employee_report_excel_fbv(request)


class PositionReportPDFCBV(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return position_report_pdf_fbv(request)


class PositionReportExcelCBV(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return position_report_excel_fbv(request)


# ==============================================================================
# 3. PUBLIC ROUTER DISPATCHERS
# ==============================================================================

def get_view_mode(request):
    return request.session.get('view_mode', 'fbv')


@login_required
def home_view(request):
    if get_view_mode(request) == 'cbv':
        return HomeCBV.as_view()(request)
    return home_view_fbv(request)


@login_required
def position_list(request):
    if get_view_mode(request) == 'cbv':
        return PositionListCBV.as_view()(request)
    return position_list_fbv(request)


@login_required
def position_create(request):
    if get_view_mode(request) == 'cbv':
        return PositionCreateCBV.as_view()(request)
    return position_create_fbv(request)


@login_required
def position_edit(request, pk):
    if get_view_mode(request) == 'cbv':
        return PositionEditCBV.as_view()(request, pk=pk)
    return position_edit_fbv(request, pk=pk)


@login_required
def position_detail(request, pk):
    if get_view_mode(request) == 'cbv':
        return PositionDetailCBV.as_view()(request, pk=pk)
    return position_detail_fbv(request, pk=pk)


@login_required
def position_delete(request, pk):
    if get_view_mode(request) == 'cbv':
        return PositionDeleteCBV.as_view()(request, pk=pk)
    return position_delete_fbv(request, pk=pk)


@login_required
def employee_list(request):
    if get_view_mode(request) == 'cbv':
        return EmployeeListCBV.as_view()(request)
    return employee_list_fbv(request)


@login_required
def employee_create(request):
    if get_view_mode(request) == 'cbv':
        return EmployeeCreateCBV.as_view()(request)
    return employee_create_fbv(request)


@login_required
def employee_edit(request, pk):
    if get_view_mode(request) == 'cbv':
        return EmployeeEditCBV.as_view()(request, pk=pk)
    return employee_edit_fbv(request, pk=pk)


@login_required
def employee_detail(request, pk):
    if get_view_mode(request) == 'cbv':
        return EmployeeDetailCBV.as_view()(request, pk=pk)
    return employee_detail_fbv(request, pk=pk)


@login_required
def employee_delete(request, pk):
    if get_view_mode(request) == 'cbv':
        return EmployeeDeleteCBV.as_view()(request, pk=pk)
    return employee_delete_fbv(request, pk=pk)


@login_required
def employee_report_pdf(request):
    if get_view_mode(request) == 'cbv':
        return EmployeeReportPDFCBV.as_view()(request)
    return employee_report_pdf_fbv(request)


@login_required
def employee_report_excel(request):
    if get_view_mode(request) == 'cbv':
        return EmployeeReportExcelCBV.as_view()(request)
    return employee_report_excel_fbv(request)


@login_required
def position_report_pdf(request):
    if get_view_mode(request) == 'cbv':
        return PositionReportPDFCBV.as_view()(request)
    return position_report_pdf_fbv(request)


@login_required
def position_report_excel(request):
    if get_view_mode(request) == 'cbv':
        return PositionReportExcelCBV.as_view()(request)
    return position_report_excel_fbv(request)
